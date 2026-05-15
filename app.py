"""
Анализ применения МП Инспектор — Streamlit-версия
Для развёртывания на Streamlit Community Cloud (GitHub)
"""

import io
import math
import re
from collections import defaultdict
from datetime import date, datetime

import openpyxl
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

# ================== Конфигурация ==================
SHEET_NAME = "Детализация МП Инспектор"

COLUMN_KEYWORDS = {
    "subjekt": ["субъект рф"],
    "podrazd": ["подразделение"],
    "vid_nadzora": ["вид надзора"],
    "nom_knm": ["номер кнм"],
    "vid": ["вид"],
    "status": ["статус кнм"],
    "narusheniya": ["нарушения выявлены"],
    "proverka_ogv": ["проверка огв/омсу"],
    "knd": ["кнд"],
    "ssylki": ["ссылки на файлы"],
    "date_act": ["дата составления акта о результате кнм", "дата составления акта"],
    "tip_prof_vizita": ["тип проф. визита", "тип профилактического визита"],
    "s_vks": ["с вкс", "вкс"],
}


# ================== Вспомогательные функции ==================

def normalize_str(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def find_column_index(headers, possible_names):
    headers_norm = [normalize_str(h) for h in headers]
    possible_norm = [normalize_str(n) for n in possible_names]
    for idx, norm in enumerate(headers_norm):
        if norm in possible_norm:
            return idx
    for idx, norm in enumerate(headers_norm):
        for pn in possible_norm:
            if pn in norm:
                return idx
    return None


def has_mp4_links(value):
    if value is None:
        return False
    return ".mp4" in str(value).lower()


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def load_data(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Лист '{SHEET_NAME}' не найден в файле.")

    ws = wb[SHEET_NAME]
    headers = [cell.value if cell.value else "" for cell in ws[1]]

    col_indices = {}
    log_lines = []
    missing = []

    for key, possible_names in COLUMN_KEYWORDS.items():
        idx = find_column_index(headers, possible_names)
        if idx is None:
            if key == "podrazd":
                if len(headers) > 17:
                    idx = 17
                    log_lines.append(f"⚠️ Столбец 'podrazd' не найден по имени — используем позицию 18 (индекс 17)")
                else:
                    missing.append(f"'podrazd' (не найден по имени и отсутствует столбец 18)")
            else:
                missing.append(f"'{key}' (искали: {possible_names})")
        else:
            log_lines.append(f"✅ Столбец **{key}** → `{headers[idx]}` (индекс {idx})")
        col_indices[key] = idx

    if missing:
        raise ValueError("Не найдены обязательные столбцы:\n• " + "\n• ".join(missing))

    data = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if not all(cell is None for cell in row)
    ]
    log_lines.append(f"📊 Загружено строк: **{len(data)}**")
    return data, col_indices, headers, log_lines


def filter_by_date(data, col_idx, date_from, date_to):
    date_col = col_idx["date_act"]
    filtered, skipped_range, skipped_invalid = [], 0, 0
    for row in data:
        parsed = parse_date(row[date_col])
        if parsed is None:
            skipped_invalid += 1
            continue
        if date_from <= parsed <= date_to:
            filtered.append(row)
        else:
            skipped_range += 1
    return filtered, skipped_range, skipped_invalid


def build_reason(vks_str, expected_value, has_links, raw_value, has_mp4=None):
    reasons = []
    if vks_str != expected_value:
        reasons.append(f"С ВКС ≠ '{expected_value}': {raw_value}")
    if not has_links:
        reasons.append("Ссылки пустые")
    if has_mp4 is not None and not has_mp4:
        reasons.append("Нет ссылок .mp4")
    return "; ".join(reasons)


def append_unique(detail, seen, key, unique_key, row):
    if unique_key not in seen[key]:
        seen[key].add(unique_key)
        detail[key].append(row)


def calculate_all_metrics(data, col_idx):
    subj_col = col_idx["subjekt"]
    podrazd_col = col_idx["podrazd"]
    knm_col = col_idx["nom_knm"]
    vid_col = col_idx["vid"]
    status_col = col_idx["status"]
    proverka_col = col_idx["proverka_ogv"]
    vid_nadzora_col = col_idx["vid_nadzora"]
    knd_col = col_idx["knd"]
    nar_col = col_idx["narusheniya"]
    vks_col = col_idx["s_vks"]
    ssylki_col = col_idx["ssylki"]
    mp4_col = 45

    allowed_vids = {"", "выездная проверка", "рейдовый осмотр", "инспекционный визит"}

    knm_info = {}
    subj_of = {}
    seen = {key: set() for key in ("vks_denom", "vks_num", "och_denom", "och_num", "och_nar_denom", "och_nar_num")}
    detail = {key: [] for key in seen}
    denom_rows_vks = {}
    denom_rows_och = {}
    rejected_vks = []
    rejected_och = []

    for row in data:
        reasons_base = []
        if normalize_str(row[status_col]) != "завершена":
            reasons_base.append(f"Статус: {row[status_col]}")
        if normalize_str(row[proverka_col]) != "нет":
            reasons_base.append(f"Проверка ОГВ: {row[proverka_col]}")
        if normalize_str(row[vid_nadzora_col]) == "гнго":
            reasons_base.append("Вид надзора: ГНГО")

        subj = row[subj_col]
        podrazd = row[podrazd_col]
        knm = row[knm_col]

        if not subj or not knm:
            reasons_base.append("Пустой субъект или номер КНМ")

        if reasons_base:
            reason_text = "; ".join(reasons_base)
            rejected_vks.append(tuple(row) + (reason_text,))
            rejected_och.append(tuple(row) + (reason_text,))
            continue

        pod_key = str(podrazd).strip() if podrazd else "Не указано"
        if pod_key not in subj_of:
            subj_of[pod_key] = str(subj).strip() if subj else ""

        vid_val = normalize_str(row[vid_col]) if row[vid_col] else ""
        knd_str = normalize_str(row[knd_col]) if row[knd_col] else ""
        nar_str = normalize_str(row[nar_col]) if row[nar_col] else ""
        vks_str = normalize_str(row[vks_col]) if row[vks_col] else ""
        ssylki_ok = row[ssylki_col] is not None and str(row[ssylki_col]).strip() != ""
        mp4_ok = has_mp4_links(row[mp4_col]) if len(row) > mp4_col else False
        sk = knm

        if vid_val in allowed_vids:
            append_unique(detail, seen, "vks_denom", sk, row)
            if vks_str == "да" and ssylki_ok and mp4_ok:
                append_unique(detail, seen, "vks_num", sk, row)
            if sk not in denom_rows_vks:
                denom_rows_vks[sk] = (row, build_reason(vks_str, "да", ssylki_ok, row[vks_col], mp4_ok))
        else:
            rejected_vks.append(tuple(row) + (f"Вид КНМ не входит в список ВКС: {row[vid_col]}",))

        if vid_val in allowed_vids:
            if "осмотр" in knd_str:
                append_unique(detail, seen, "och_denom", sk, row)
                if vks_str == "нет" and ssylki_ok:
                    append_unique(detail, seen, "och_num", sk, row)
                if sk not in denom_rows_och:
                    denom_rows_och[sk] = (row, build_reason(vks_str, "нет", ssylki_ok, row[vks_col]))
                if nar_str == "да":
                    append_unique(detail, seen, "och_nar_denom", sk, row)
                    if vks_str == "нет" and ssylki_ok:
                        append_unique(detail, seen, "och_nar_num", sk, row)
            else:
                rejected_och.append(tuple(row) + (f"КНД не содержит 'осмотр': {row[knd_col]}",))
        else:
            rejected_och.append(tuple(row) + (f"Вид КНМ не входит в список очных: {row[vid_col]}",))

        if knm not in knm_info:
            knm_info[knm] = {
                "podr": pod_key,
                "vks_denom": False, "vks_num": False,
                "och_denom": False, "och_num": False, "och_nar": False,
            }

        info = knm_info[knm]
        if vid_val in allowed_vids:
            info["vks_denom"] = True
            if vks_str == "да" and ssylki_ok and mp4_ok:
                info["vks_num"] = True
        if vid_val in allowed_vids and "осмотр" in knd_str:
            info["och_denom"] = True
            if vks_str == "нет" and ssylki_ok:
                info["och_num"] = True
            if nar_str == "да":
                info["och_nar"] = True

    metrics = defaultdict(lambda: [set(), set(), set(), set(), set()])
    for knm, info in knm_info.items():
        podr = info["podr"]
        if info["vks_denom"]: metrics[podr][0].add(knm)
        if info["vks_num"]:   metrics[podr][1].add(knm)
        if info["och_denom"]: metrics[podr][2].add(knm)
        if info["och_num"]:   metrics[podr][3].add(knm)
        if info["och_nar"]:   metrics[podr][4].add(knm)

    denom_not_num_vks = [
        tuple(row) + (reason,)
        for sk, (row, reason) in denom_rows_vks.items()
        if sk not in seen["vks_num"]
    ]
    denom_not_num_och = [
        tuple(row) + (reason,)
        for sk, (row, reason) in denom_rows_och.items()
        if sk not in seen["och_num"]
    ]

    return metrics, subj_of, detail, rejected_vks, rejected_och, denom_not_num_vks, denom_not_num_och


def build_report_data(metrics, subj_of):
    result = []
    for pod, sets in metrics.items():
        result.append({
            "Подразделение": pod,
            "Субъект": subj_of.get(pod, ""),
            "total_vks": len(sets[0]),
            "prim_vks": len(sets[1]),
            "total_och": len(sets[2]),
            "prim_och": len(sets[3]),
            "total_och_nar": len(sets[4]),
        })
    return result


def save_to_excel(report_data, headers_orig, detail, rejected_vks, rejected_och, denom_not_num_vks, denom_not_num_och):
    wb = openpyxl.Workbook()

    def make_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, end_color=hex_color)

    def style_row(ws, row_num, fill_color, bold=True, font_color="FFFFFF"):
        fill = make_fill(fill_color)
        font = Font(bold=bold, color=font_color, name="Arial", size=10)
        for cell in ws[row_num]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autowidth(ws):
        for col in ws.columns:
            width = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 60)

    def write_detail_sheet(title, rows, fill_color, extra_header=None):
        ws = wb.create_sheet(title=title)
        headers = list(headers_orig) + ([extra_header] if extra_header else [])
        ws.append(headers)
        style_row(ws, 1, fill_color)
        ws.row_dimensions[1].height = 40
        for row in rows:
            ws.append(list(row))
        autowidth(ws)

    ws = wb.active
    ws.title = "Итоги по подразделениям"

    ws.append(["", "", "ВКС", "", "ОЧНЫЕ", ""])
    ws.merge_cells("C1:D1")
    ws.merge_cells("E1:F1")

    blue = "4472C4"
    itog = "1F3864"

    for cell in ws[1]:
        cell.fill = make_fill(blue)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.append([
        "№ п/п", "Подразделение",
        "Всего в ААС КНД", "Всего в МП Инспектор (доля, %)",
        "Всего в ААС КНД\n(из них с нарушениями)", "Всего в МП Инспектор (доля, %)",
    ])
    style_row(ws, 2, blue)
    ws.row_dimensions[2].height = 45

    col_widths = [7, 45, 20, 30, 28, 30]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    def och_pct(item):
        return item["prim_och"] / item["total_och"] if item["total_och"] > 0 else 0.0

    report_sorted = sorted(report_data, key=och_pct)

    def fmt_vks(prim, total):
        pct = prim / total * 100 if total > 0 else 0.0
        return f"{prim} ({pct:.2f}%)"

    def fmt_och_denom(total, nar):
        return f"{total} ({nar})"

    def fmt_och_num(prim, total):
        pct = prim / total * 100 if total > 0 else 0.0
        return f"{prim} ({pct:.2f}%)"

    row_number = 1
    for item in report_sorted:
        ws.append([
            row_number, item["Подразделение"],
            item["total_vks"], fmt_vks(item["prim_vks"], item["total_vks"]),
            fmt_och_denom(item["total_och"], item["total_och_nar"]),
            fmt_och_num(item["prim_och"], item["total_och"]),
        ])
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=10)
        ws.cell(current_row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if row_number % 2 == 0:
            for cell in ws[current_row]:
                cell.fill = make_fill("DCE6F1")

        pod_text = str(item["Подразделение"]) if item["Подразделение"] else ""
        col2_width = 43
        line_height_pt = 14
        padding_pt = 6
        lines_needed = math.ceil(len(pod_text) / col2_width) if pod_text else 1
        ws.row_dimensions[current_row].height = max(lines_needed * line_height_pt + padding_pt, 20)
        row_number += 1

    total_vks = sum(item["total_vks"] for item in report_data)
    prim_vks = sum(item["prim_vks"] for item in report_data)
    total_och = sum(item["total_och"] for item in report_data)
    prim_och = sum(item["prim_och"] for item in report_data)
    total_nar = sum(item["total_och_nar"] for item in report_data)

    ws.append([
        "", "Итог по субъекту",
        total_vks, fmt_vks(prim_vks, total_vks),
        fmt_och_denom(total_och, total_nar), fmt_och_num(prim_och, total_och),
    ])

    itog_row = ws.max_row
    for cell in ws[itog_row]:
        cell.fill = make_fill(itog)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(itog_row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[itog_row].height = 24

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=itog_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border

    write_detail_sheet("ВКС всего", detail["vks_denom"], "538135")
    write_detail_sheet("ВКС с МП", detail["vks_num"], "C55A11")
    write_detail_sheet("ВКС без МП", denom_not_num_vks, "7030A0", "Причина")
    write_detail_sheet("ВКС — Отсеянные", rejected_vks, "C00000", "Причина отклонения")
    write_detail_sheet("Очные всего", detail["och_denom"], "538135")
    write_detail_sheet("Очные с МП", detail["och_num"], "C55A11")
    write_detail_sheet("Очные без МП", denom_not_num_och, "7030A0", "Причина")
    write_detail_sheet("Очные всего (нар.)", detail["och_nar_denom"], "375623")
    write_detail_sheet("Очные с МП (нар.)", detail["och_nar_num"], "843C0C")
    write_detail_sheet("Очные — Отсеянные", rejected_och, "C00000", "Причина отклонения")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ================== Streamlit UI ==================

def main():
    st.set_page_config(
        page_title="МП Инспектор — Анализ",
        page_icon="🔥",
        layout="centered",
    )

    st.title("🔥 Анализ применения МП Инспектор")
    st.caption("Итоги по подразделениям · ОНД и ПР · МЧС России")

    st.divider()

    # 1. Загрузка файла
    st.subheader("1. Выгрузка из АИС КНД")
    uploaded = st.file_uploader(
        "Выберите файл выгрузки (.xlsx / .xlsm)",
        type=["xlsx", "xlsm"],
        help=f"Файл должен содержать лист «{SHEET_NAME}»",
    )

    if not uploaded:
        st.info("Загрузите файл выгрузки для начала работы.")
        return

    # 2. Выбор периода
    st.subheader("2. Период анализа")
    now = datetime.now()
    col1, col2 = st.columns(2)
    with col1:
        date_from = st.date_input(
            "Дата начала периода",
            value=date(now.year, 1, 1),
            format="DD.MM.YYYY",
        )
    with col2:
        date_to = st.date_input(
            "Дата окончания периода",
            value=date(now.year, now.month, now.day),
            format="DD.MM.YYYY",
        )

    if date_from > date_to:
        st.warning("⚠️ Дата начала позже даты окончания — они будут переставлены местами.")
        date_from, date_to = date_to, date_from

    st.caption(f"Анализируемый период: **{date_from.strftime('%d.%m.%Y')}** — **{date_to.strftime('%d.%m.%Y')}**")

    # 3. Кнопка запуска
    st.divider()
    if not st.button("▶ Сформировать отчёт", type="primary", use_container_width=True):
        return

    # Загрузка и обработка
    with st.spinner("Загрузка данных..."):
        try:
            file_bytes = uploaded.read()
            data, col_idx, headers_orig, log_lines = load_data(file_bytes)
        except Exception as e:
            st.error(f"❌ Ошибка при загрузке файла:\n{e}")
            return

    with st.expander("📋 Лог определения столбцов", expanded=False):
        for line in log_lines:
            st.markdown(line)

    with st.spinner("Фильтрация по дате..."):
        data_filtered, skipped_range, skipped_invalid = filter_by_date(data, col_idx, date_from, date_to)

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Строк в периоде", len(data_filtered))
    col_b.metric("Строк вне периода", skipped_range)
    col_c.metric("Строк с нераспознанной датой", skipped_invalid)

    if len(data_filtered) == 0:
        st.warning("⚠️ Нет данных за выбранный период.")
        return

    with st.spinner("Расчёт показателей..."):
        metrics, subj_of, detail, rej_vks, rej_och, dnn_vks, dnn_och = calculate_all_metrics(data_filtered, col_idx)
        report_data = build_report_data(metrics, subj_of)

    st.success(f"✅ Обработано подразделений: **{len(metrics)}**")

    # Сводная таблица в интерфейсе
    st.subheader("📊 Сводная таблица")

    def och_pct(item):
        return item["prim_och"] / item["total_och"] if item["total_och"] > 0 else 0.0

    report_sorted = sorted(report_data, key=och_pct)

    table_rows = []
    for i, item in enumerate(report_sorted, 1):
        t_vks = item["total_vks"]
        p_vks = item["prim_vks"]
        t_och = item["total_och"]
        p_och = item["prim_och"]
        nar   = item["total_och_nar"]
        pct_vks = p_vks / t_vks * 100 if t_vks else 0
        pct_och = p_och / t_och * 100 if t_och else 0
        table_rows.append({
            "№": i,
            "Подразделение": item["Подразделение"],
            "ВКС всего": t_vks,
            "ВКС МП (%)": f"{p_vks} ({pct_vks:.1f}%)",
            "Очные (нар.)": f"{t_och} ({nar})",
            "Очные МП (%)": f"{p_och} ({pct_och:.1f}%)",
        })

    import pandas as pd
    df = pd.DataFrame(table_rows).set_index("№")
    st.dataframe(df, use_container_width=True)

    # Статистика детализации
    st.subheader("📈 Статистика детализации")
    stat_cols = st.columns(2)
    with stat_cols[0]:
        st.markdown("**ВКС**")
        st.write(f"• Всего КНМ: {len(detail['vks_denom'])}")
        st.write(f"• С МП: {len(detail['vks_num'])}")
        st.write(f"• Без МП: {len(dnn_vks)}")
        st.write(f"• Отсеянных: {len(rej_vks)}")
    with stat_cols[1]:
        st.markdown("**Очные**")
        st.write(f"• Всего КНМ: {len(detail['och_denom'])}")
        st.write(f"• С МП: {len(detail['och_num'])}")
        st.write(f"• Без МП: {len(dnn_och)}")
        st.write(f"• Очных (нар.) всего: {len(detail['och_nar_denom'])}")
        st.write(f"• Очных (нар.) с МП: {len(detail['och_nar_num'])}")
        st.write(f"• Отсеянных: {len(rej_och)}")

    # Формирование и скачивание файла
    st.divider()
    with st.spinner("Формирование Excel-файла..."):
        try:
            xlsx_bytes = save_to_excel(
                report_data, headers_orig,
                detail, rej_vks, rej_och, dnn_vks, dnn_och,
            )
        except Exception as e:
            st.error(f"❌ Ошибка при формировании файла: {e}")
            return

    filename = f"результат_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="⬇️ Скачать результат (.xlsx)",
        data=xlsx_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
